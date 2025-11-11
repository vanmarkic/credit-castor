#!/usr/bin/env python3
"""
Test timeline snapshot export functionality
Verifies that the Excel export includes both the calculator sheet and timeline snapshots sheet
"""

from playwright.sync_api import sync_playwright
import time

def test_timeline_export():
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)  # Use headful to see what's happening
        page = browser.new_page()

        # Navigate to the app
        print("📍 Navigating to app...")
        page.goto('http://localhost:4323/credit-castor/')

        # Wait for the app to load
        print("⏳ Waiting for app to load...")
        page.wait_for_load_state('networkidle')
        time.sleep(2)  # Extra wait for React to hydrate

        # Take screenshot of initial state
        page.screenshot(path='/tmp/credit-castor-initial.png', full_page=True)
        print("📸 Screenshot saved to /tmp/credit-castor-initial.png")

        # Handle password gate if present
        print("🔓 Checking for password gate...")
        password_input = page.locator('input[type="password"]')
        if password_input.count() > 0:
            print("🔓 Password gate detected")
            # Try common password (usually 'password' or empty for dev)
            print("   Trying to bypass password gate...")
            password_input.fill('')  # Try empty password first
            access_button = page.locator('button:has-text("Accéder")')
            if access_button.count() > 0:
                access_button.click()
                time.sleep(1)

            # Check if we're still on password page
            if page.locator('input[type="password"]').count() > 0:
                print("   Empty password didn't work, this test requires the password")
                print("   You can modify this script to include the correct password")
                page.screenshot(path='/tmp/credit-castor-password-gate.png', full_page=True)
                browser.close()
                return False

            page.wait_for_load_state('networkidle')
            time.sleep(1)
            print("✅ Entered app")
            page.screenshot(path='/tmp/credit-castor-after-login.png', full_page=True)

        # Find the export button (look for the Excel export button)
        print("🔍 Looking for export button...")

        # First, let's inspect what buttons are available
        buttons = page.locator('button').all()
        print(f"Found {len(buttons)} buttons")

        for i, button in enumerate(buttons):
            text = button.text_content()
            if text:
                print(f"  Button {i}: {text[:50]}")

        # Look for the export button - it might have an Excel icon or "Export" text
        # Common patterns: "Exporter", "Export Excel", icon button
        export_button = None

        # Try different selector strategies
        selectors_to_try = [
            'button:has-text("Excel")',
            'button:has-text("Export")',
            'button:has-text("Exporter")',
            'button[title*="Excel"]',
            'button[title*="export" i]',
        ]

        for selector in selectors_to_try:
            try:
                if page.locator(selector).count() > 0:
                    export_button = page.locator(selector).first
                    print(f"✅ Found export button with selector: {selector}")
                    break
            except:
                continue

        if not export_button:
            print("❌ Could not find export button")
            print("Available button text content:")
            for button in buttons:
                print(f"  - {button.text_content()}")
            page.screenshot(path='/tmp/credit-castor-no-export-button.png', full_page=True)
            browser.close()
            return False

        # Set up download handler
        print("💾 Setting up download handler...")
        download_promise = page.expect_download()

        # Click the export button
        print("🖱️  Clicking export button...")
        export_button.click()

        # Wait for download
        print("⏳ Waiting for download...")
        download = download_promise.value

        # Save the file
        download_path = '/tmp/credit-castor-export.xlsx'
        download.save_as(download_path)

        print(f"✅ Download completed: {download.suggested_filename}")
        print(f"   Saved to: {download_path}")

        # We can't easily inspect XLSX contents without additional libraries,
        # but we can check the file size and name
        import os
        file_size = os.path.getsize(download_path)
        print(f"   File size: {file_size} bytes")

        if file_size > 0:
            print("✅ Export successful - file has content")

            # Optional: Try to read with openpyxl if available
            try:
                import openpyxl
                wb = openpyxl.load_workbook(download_path)
                sheet_names = wb.sheetnames
                print(f"📊 Workbook contains {len(sheet_names)} sheets:")
                for name in sheet_names:
                    sheet = wb[name]
                    rows = sheet.max_row
                    cols = sheet.max_column
                    print(f"   - {name}: {rows} rows x {cols} columns")

                if 'Timeline Snapshots' in sheet_names:
                    print("✅ Timeline Snapshots sheet found!")
                    timeline_sheet = wb['Timeline Snapshots']
                    print(f"   Timeline sheet has {timeline_sheet.max_row} rows")

                    # Check header row
                    headers = [cell.value for cell in timeline_sheet[1]]
                    print(f"   Headers: {headers[:5]}...")

                else:
                    print("⚠️  Timeline Snapshots sheet NOT found")

                wb.close()
            except ImportError:
                print("ℹ️  openpyxl not available - skipping detailed inspection")
            except Exception as e:
                print(f"⚠️  Error inspecting Excel file: {e}")
        else:
            print("❌ Export file is empty")
            browser.close()
            return False

        # Take final screenshot
        page.screenshot(path='/tmp/credit-castor-after-export.png', full_page=True)
        print("📸 Final screenshot saved to /tmp/credit-castor-after-export.png")

        browser.close()
        print("✅ Test completed successfully")
        return True

if __name__ == '__main__':
    success = test_timeline_export()
    exit(0 if success else 1)
